#libraries that are used in this script
import sys; print(sys.version)  #build system
import os
from datetime import datetime, date; print("CurrentDate:", datetime.now(), "\n\n") #date stamp
print("---------------------Welcome to PIP Generator---------------------\n")
import xml.sax #.xml parser
import openpyxl #create a workbook
from openpyxl.workbook import Workbook #inserting line in workbook
from openpyxl import load_workbook
from openpyxl.worksheet import *
from openpyxl.cell import Cell
from openpyxl.utils.cell import get_column_letter, column_index_from_string, coordinate_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors
import json





mcu = ["PSOC", "NRF51", "RENESAS", "XILOG"];

equipment = ["TREADMILL_DEVICE", "INCLINE_TRAINER_DEVICE", "ELLIPTICAL_DEVICE",
             "EXERCISE_BIKE_DEVICE", "SPIN_BIKE_DEVICE", "VERTICAL_ELLIPTICAL_DEVICE",
             "FITNESS_WT_MACH_DEVICE", "FREE_STRIDER_DEVICE", "ROWER_DEVICE"];

display = ["NO_DISPLAY", "SISP_DISPLAY", "AW_DISPLAY", "AW_DESK", "RICKENBACKER_DISPLAY",#4
           "RICKENBACKER_MINI_DISPLAY", "PM210_DISPLAY_PAUL_REED_SMITH", "PM210_DISPLAY_GIBSON_NORDIC",#7
           "POWER_RING_DISPLAY", "POWER_RING_DISPLAY_LARGE", "AERO_POWER_RING_DISPLAY","TTW_DISPLAY",#11
           "LED_900", "CD_DISPLAY", "FENDER_DISPLAY", "OFENDER_DISPLAY", "OP_DISPLAY", "IF17_DISPLAY",#17
           "IF17_VERT_DISPLAY", "MOON_DISPLAY", 'AW_AEROBIC_DISPLAY'];#20

tablet = ["RoyalWolf", "Legacy", "Legacy_FreeMotion", "Argon", "Helium"]

controller = ["TREAD_TACH", "MC1650LS_2W", "MC1618DLS", "MC1618DLS_JST", "MC1648DLS", "MC1648DLS_JST", "MC1618IHS",#6
              "MC1648IHS", "MC1618IHB", "MC1648IHB", "MC2100LT_12", "MC2100LTS_30", "MC2100LTS_50W", "MC2100ELS_18W_2Y",#13
              "MC2100ELS_50W_2Y", "MC5100DTS_18W", "MC5100DTS_50W", "MC5100DTS3_50W", "MC5100EDS_50W_V1",#18
              "MC5100EDS3_50W_V1", "MC5150HCL", "OLYMPUS_V1", "TIMPANOGOS", "RHYMEBUS_RM", "DELTA GOES HERE", "BASIC_BIKE",#25
              "PB_INC_18W", "PB_INC_50W", "PB_INC_485_18W", "PB_INC_485_50W", "PB_PPI_485_18W", "PB_PPI_485_48W",#31
              "PBCLA_2X50W_75W", "PBCLA_FM_32_12V", "KIWI"]; #0-18 home units, 21-24 club units, 25-33 bike and elliptical, 34 KIWI

voltage = ["0", "6", "9", "12"];
resistance = ["ADC"];
distance = ["NONE", "LoResTach", "LoResAccelerometerTach", "PowerBoardAccelerometer", "LoResNewAccelerometer"];

maintainance = ["CALIBRATE", "TOGGLE_DISPLAY_ON_OFF", "KEYCODE_TEST"]; #tablet consoles will be different

keycode = ["ON_RESET", "SETTINGS", "POWER_ON_OFF", "INCLINE_UP"]

ble = ["NO_PROFILE_SUPPORTED", "WEIGHT_MACHINE_SUPPORTED", "SPEED_RING_SUPPORTED", "FILE_SYSTEM_SUPPORTED",
       "FITPRO_SYSTEM_SUPPORTED", "HEART_RATE_SENSOR_SUPPORTED"];

pulse = ["Hand", "Thumb", "Chest", "Ant", "BLE", "nanoHand", "Priority"];

fan = ["TwoWire", "ThreeWire", "twoAndThree"];

audio = ["None", "PC", "BrainBoard", "MP3", "iPod", "Headphones","BLE", "TV", "MYE"]; #audio functions

header = ["void for function call", "RecordID", "Link to Parts - Console Pics", "Link to Parts - Consoles",
          "Part Number","Description", "ProductType", "Display Type", "Equipment Needed", "Setup", "BLE SETUP",
          "Part Type Pull", "PIP", "Video of Inspection", "EQF1259 Config", "BLE Setup Image 1", "BLE Setup Image 2",
          "Setup Image 1", "Setup Image 2", "ESD Bag", "Cosmetics", "Warning Label", "Warning Label Picture",
          "Cosmetic, Color, and Texture", "Picture of Console", "Soldering Quality", "Power up", "Safety Key",
          "BLE Test", "BLE Test Image 1", "BLE Test Image 2", "Software Version", "Test Mode Image", "Incline Calibration",
          "Display Test", "Display Test Image", "Button Test", "Button Test Image", "Drive Motor Output Test",
          "Tach Input Test", "Resistance Motor Test", "Incline Motor Test", "USB port test", "Hand Pulse Test",
          "Chest Pulse Test", "Fan Test", "Audio Test", "TV Test", "Upright Motor Test", "Finish Test",
          "Last User", "Created By User", "Updated", "Created"];



#EQF1259 source and to
app = load_workbook(filename = "ConsoleProcedures.xlsx")
#Read sheets
eng = app["english"] #PIP database
chi = app["chinese"]


#Create standalone
mywb = openpyxl.Workbook()
mywb.remove(mywb["Sheet"]) #removes default sheet title
mywb.create_sheet(index=0, title="Parts - Consoles") #adds new sheet with title
mywb.create_sheet(index=1, title="revision")

sta = mywb["Parts - Consoles"] #pip
for i in range(1, 54):
    _ = sta.cell(column = i, row = 1, value = header[i]) #write headers

for col in sta.columns:
     max_length = 0
     column = col[0].column # Get the column name
     for cell in col:

         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = 100
     sta.column_dimensions[column].width = adjusted_width #adjust multi column width that contains string


rev = mywb["revision"] #rev
#Preliminary
rev["A1"] = "V.0"
rev["B1"] = "20171005"
rev["C1"] = "Order from chaos"
rev["D1"] = "id-0"

rev["A2"] = "V.1"
rev["B2"] = "20171205"
rev["C2"] = "Created arrays for variables"
rev["D2"] = "id-0"

rev["A3"] = "V.2"
rev["B3"] = "20171206"
rev["C3"] = "Added time stamping"
rev["D3"] = "id-0"

rev["A4"] = "V.3"
rev["B4"] = "20171207"
rev["C4"] = "Functions on test x.xml"
rev["D4"] = "id-0"

rev["A5"] = "V.4"
rev["B5"] = "20171227"
rev["C5"] = "Converted to openpyxl"
rev["D5"] = "id-0"

rev["A6"] = "V.5"
rev["B6"] = "20180112"
rev["C6"] = "Killed trackvia full table"
rev["D6"] = "id-0"

rev["A7"] = "V.6"
rev["B7"] = "20180123"
rev["C7"] = "Added chinese and auto launch for PIP"
rev["D7"] = "id-0"

rev["A8"] = "V.7"
rev["B8"] = "20180223"
rev["C8"] = "About done"
rev["D8"] = "id-0"

rev["A9"] = "V.8"
rev["B9"] = "20180228"
rev["C9"] = "Migrated to Python3.6.4"
rev["D9"] = "id-0"

rev["A10"] = "V.9"
rev["B10"] = "20180307"
rev["C10"] = "Test build with cx_Freeze5.1.1"
rev["D10"] = "id-0"

#Release
rev["A11"] = "V.10"
rev["B11"] = "20180327"
rev["C11"] = "Added pause to verify no errors"
rev["D11"] = "id-0"

rev["A12"] = "V.11"
rev["B12"] = "20180402"
rev["C12"] = "Working on Basic_Bike callouts"
rev["D12"] = "id-0"

rev["A13"] = "V.12"
rev["B13"] = "20180418"
rev["C13"] = "BLE finish steps work again"
rev["D13"] = "id-0"

rev["A14"] = "V.13"
rev["B14"] = "20181015"
rev["C14"] = "FreeMotion Legacy tablet /w MYE"
rev["D14"] = "id-0"

rev["A15"] = "V.14"
rev["B15"] = "20181217"
rev["C15"] = "Working on BASIC_BIKE"
rev["D15"] = "id-0"

rev["A16"] = "V.15"
rev["B16"] = "20190221"
rev["C16"] = "Added the VM callouts"
rev["D16"] = "id-0"

rev["A17"] = "V.16"
rev["B17"] = "20190222" 
rev["C17"] = "Issue with Rhymebus callouts fixed"
rev["D17"] = "id-0"

rev["A18"] = "V.17"
rev["B18"] = "20190228"
rev["C18"] = "Legacy tablet steps added"
rev["D18"] = "id-0"

rev["A19"] = "V.18"
rev["B19"] = "20190524"
rev["C19"] = "Added _JST controllers"
rev["D19"] = "id-0"

rev["A20"] = "V.19"
rev["B20"] = "20190716"
rev["C20"] = "Console name is now output file + _PIP.xlsx"
rev["D20"] = "id-0"

rev["A21"] = "V.20"
rev["B21"] = "20190723"
rev["C21"] = "Fixed a BLE bike maintenance issue"
rev["D21"] = "id-0"

rev["A22"] = "V.21"
rev["B22"] = "20190923"
rev["C22"] = "Added array features to some functions"
rev["D22"] = "Ross Bunnell and id-0"

rev["A23"] = "V.22"
rev["B23"] = "20191001"
rev["C23"] = "added .json generator"
rev["D23"] = "Ross Bunnell and id-0"

rev["A24"] = "V.23"
rev["B24"] = "20200921"
rev["C24"] = "Added the KIWI board"
rev["D24"] = "id-0"

rev["A25"] = "V.x"
rev["B25"] = str(datetime.now())
rev["C25"] = "Fiat justitia ruat caelum"
rev["D25"] = "id-0"

print (rev["a24"].value, rev["b24"].value, rev["c24"].value, rev["d24"].value, "\n\n")


for col in rev.columns:
     max_length = 0
     column = col[0].column # Get the column name
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = (max_length + 2) * 1.2 #add defined width value
     rev.column_dimensions[column].width = adjusted_width #adjust multi column width that contains string

temp = ""

#Parse .xml file
xmlDict = {}
pulsearray = []
audioarray = []
keycodearray = []
keycodecatarray = []
class Device( xml.sax.ContentHandler ):
   def __init__(self):

      self.CurrentData = ""
      self.consolePartNumber = ""
      self.mcuChipName = ""
      self.equipmentType = ""
      self.systemUnitType = ""
      self.BuildModelString = ""
      self.TypeName = ""
      self.PowerBoard = ""
      self.ConsoleVoltage = ""
      self.GradeProtocol = ""
      self.TabletProtocol = ""
      self.MaintenanceConfigFunction = ""
      self.KeyCodeName = ""
      self.PulseDriverItem = ""
      self.FanProtocol = ""
      self.AudioSrcItem = ""
      self.DistanceDriver = ""
      self.ResistanceDriver = ""
      self.KeyCodeCategory = ""
      



   # Call when an element starts
   def startElement(self, tag, attributes):
      self.CurrentData = tag
      if tag == "Config":
         print ("---------System_Info---------\n")
      elif tag == "FeatureItem": #feature tag reading
         feature = attributes["xsi:type"]
         #print feature
         if feature == "TvFeature": #39818 TV test
             sta["au2"].alignment = Alignment(wrap_text = True) #MYE TV test
             sta["au2"] = chi["w4"].value + "\n\n" + eng["w4"].value
             
         
    # Call when a character is read

   def characters(self, content):
      global xmlDict
      if self.CurrentData == "consolePartNumber":
         self.consolePartNumber = content
         if  "consolePartNumber" not in xmlDict:
             xmlDict["consolePartNumber"] = self.consolePartNumber
      elif self.CurrentData == "BuildModelString":
         self.BuildModelString = content
         if  "BuildModelString" not in xmlDict:
             xmlDict["BuildModelString"] = self.BuildModelString
      elif self.CurrentData == "equipmentType":
         self.equipmentType = content
         if  "equipmentType" not in xmlDict:
             xmlDict["equipmentType"] = self.equipmentType         
      elif self.CurrentData == "TypeName":
         self.TypeName = content
         if  "TypeName" not in xmlDict:
             xmlDict["TypeName"] = self.TypeName 
      elif self.CurrentData == "mcuChipName":
         self.mcuChipName = content
         if  "mcuChipName" not in xmlDict:
             xmlDict["mcuChipName"] = self.mcuChipName 
      elif self.CurrentData == "systemUnitType":
         self.systemUnitType = content
         if  "systemUnitType" not in xmlDict:
             xmlDict["systemUnitType"] = self.systemUnitType 
      elif self.CurrentData == "PowerBoard":
         self.PowerBoard = content
         if  "PowerBoard" not in xmlDict:
             xmlDict["PowerBoard"] = self.PowerBoard 
      elif self.CurrentData == "ConsoleVoltage":
         self.ConsoleVoltage = content
         if  "ConsoleVoltage" not in xmlDict:
             xmlDict["ConsoleVoltage"] = self.ConsoleVoltage 
      elif self.CurrentData == "USBHostBoard":
         self.USBHostBoard = content
         if  "USBHostBoard" not in xmlDict:
             xmlDict["USBHostBoard"] = self.USBHostBoard 
      elif self.CurrentData == "GradeProtocol":
         self.GradeProtocol = content
         if  "GradeProtocol" not in xmlDict:
             xmlDict["GradeProtocol"] = self.GradeProtocol 
      elif self.CurrentData == "ResistanceDriver":
         self.ResistanceDriver = content
         if  "ResistanceDriver" not in xmlDict:
             xmlDict["ResistanceDriver"] = self.ResistanceDriver 
      elif self.CurrentData == "DistanceDriver":
         self.DistanceDriver = content
         if  "DistanceDriver" not in xmlDict:
             xmlDict["DistanceDriver"] = self.DistanceDriver 
      elif self.CurrentData == "TabletProtocol":
         self.TabletProtocol = content
         if  "TabletProtocol" not in xmlDict:
             xmlDict["TabletProtocol"] = self.TabletProtocol 
      elif self.CurrentData == "MaintenanceConfigFunction":
         self.MaintenanceConfigFunction = content
         if  "MaintenanceConfigFunction" not in xmlDict:
             xmlDict["MaintenanceConfigFunction"] = self.MaintenanceConfigFunction 
      elif self.CurrentData == "KeyCodeName":
         self.KeyCodeName = content
         if  "KeyCodeName" not in xmlDict:
             xmlDict["KeyCodeName"] = keycodearray
      elif self.CurrentData == "PulseDriverItem":
         self.PulseDriverItem = content
         if  "PulseDriverItem" not in xmlDict:
             xmlDict["PulseDriverItem"] = pulsearray 
      elif self.CurrentData == "FanProtocol":
         self.FanProtocol = content
         if  "FanProtocol" not in xmlDict:
             xmlDict["FanProtocol"] = self.FanProtocol 
      elif self.CurrentData == "AudioSrcItem":
         self.AudioSrcItem = content
         if  "AudioSrcItem" not in xmlDict:
             xmlDict["AudioSrcItem"] = audioarray
      elif self.CurrentData == "KeyCodeCategory":
         self.KeyCodeCategory = content
         if  "KeyCodeCategory" not in xmlDict:
             xmlDict["KeyCodeCategory"] = keycodecatarray 



   # Call when an elements ends
   def endElement(self, tag):
      global xmlDict
      if self.CurrentData == "PulseDriverItem":
         pulsearray.append(self.PulseDriverItem)
      if self.CurrentData == "AudioSrcItem":
         audioarray.append(self.AudioSrcItem)
      if self.CurrentData == "KeyCodeName":
          keycodearray.append(self.KeyCodeName)
      if self.CurrentData == "KeyCodeCategory":
          keycodecatarray.append(self.KeyCodeCategory)

         
   def BuildXml(self, CurrentData, Value):
      global temp
      if CurrentData == "consolePartNumber":
         print ("Part Number:", Value)
         sta["d2"].alignment = Alignment(wrap_text = True)
         sta["d2"] = Value
         
      elif CurrentData == "mcuChipName":
         print ("Processor:", Value)
         if Value == mcu[1]:
             sta["j2"].alignment = Alignment(wrap_text = True) #BLE setup
             sta["j2"] = chi["d2"].value + "\n\n" + eng["d2"].value
             sta["ab2"].alignment = Alignment(wrap_text = True) #BLE test
             sta["ab2"] = chi["j2"].value + "\n\n" + eng["j2"].value
             sta["ar2"].alignment = Alignment(wrap_text = True) #BLE pulse
             sta["ar2"] = chi["t5"].value + "\n\n" + eng["t5"].value
             if temp == "":
                 temp = Value

      elif CurrentData == "equipmentType":
         print ("Equipment:", Value)
         sta["f2"].alignment = Alignment(wrap_text = True)
         sta["f2"] = Value
         i = 0
         for i in range(2):
             if Value == equipment[i]:
                 sta["u2"].alignment = Alignment(wrap_text = True) #Warning label
                 sta["u2"] = chi["f2"].value + "\n\n" + eng["f2"].value
                 sta["aa2"].alignment = Alignment(wrap_text = True) #DMK
                 sta["aa2"] = chi["i2"].value + "\n\n" + eng["i2"].value
                 sta["al2"].alignment = Alignment(wrap_text = True) #Drivemotor
                 sta["al2"] = chi["o2"].value + "\n\n" + eng["o2"].value
             if temp == mcu[1]:
                 sta["ae2"].alignment = Alignment(wrap_text = True) #BLE maintainance
                 sta["ae2"] = chi["k5"].value + "\n\n" + eng["k5"].value
                 sta["ag2"].alignment = Alignment(wrap_text = True) #BLE calibrate
                 sta["ag2"] = chi["l5"].value + "\n\n" + eng["l5"].value
                 sta["ah2"].alignment = Alignment(wrap_text = True) #BLE display
                 sta["ah2"] = chi["m5"].value + "\n\n" + eng["m5"].value
                 sta["aj2"].alignment = Alignment(wrap_text = True) #BLE keycodes
                 sta["aj2"] = chi["n5"].value + "\n\n" + eng["n5"].value
                 sta["aw2"].alignment = Alignment(wrap_text = True) #Finish does not work due to location of equipmentType
                 sta["aw2"] = chi["aa5"].value + "\n\n" + eng["aa5"].value
                 
      elif CurrentData == "systemUnitType":
         print ("Unit Measure:", Value)
              
      elif CurrentData == "BuildModelString":
         print ("Part Name:", Value)
         sta["e2"].alignment = Alignment(wrap_text = True)
         sta["e2"] = Value
         
                 
      elif CurrentData == "TypeName": #----------------------------------------------Isolates tag to array items only
         if Value == display[0]:
             print ("Display:", Value)
             sta["g2"].alignment = Alignment(wrap_text = True)
             sta["g2"] = Value
         for i in range(1, 21):
             if Value == display[i]:
                 print ("Display:", Value)
                 sta["g2"].alignment = Alignment(wrap_text = True)
                 sta["g2"] = Value
                 sta["w2"].alignment = Alignment(wrap_text = True) #Cosmetic all other product
                 sta["w2"] = chi["g2"].value + "\n\n" + eng["g2"].value
         if Value == display[12]: #-------------------------------------------------------LED_900
             sta["j2"].alignment = Alignment(wrap_text = True) #BLE setup
             sta["j2"] = chi["d2"].value + "\n\n" + eng["d2"].value
             sta["ab2"].alignment = Alignment(wrap_text = True) #BLE test
             sta["ab2"] = chi["j2"].value + "\n\n" + eng["j2"].value
             sta["ae2"].alignment = Alignment(wrap_text = True) #VM maintainance
             sta["ae2"] = chi["k10"].value + "\n\n" + eng["k10"].value
             sta["ag2"].alignment = Alignment(wrap_text = True) #VM calibrate
             sta["ag2"] = chi["l10"].value + "\n\n" + eng["l10"].value
             sta["ah2"].alignment = Alignment(wrap_text = True) #VM display
             sta["ah2"] = chi["m10"].value + "\n\n" + eng["m10"].value
             sta["aj2"].alignment = Alignment(wrap_text = True) #VM keycodes
             sta["aj2"] = chi["n10"].value + "\n\n" + eng["n10"].value
             sta["aw2"].alignment = Alignment(wrap_text = True) #VM Finish
             sta["aw2"] = chi["aa4"].value + "\n\n" + eng["aa4"].value
         if Value == display[14]: #-------------------------------------------------------!OT-fendor
             sta["j2"].alignment = Alignment(wrap_text = True) #BLE setup
             sta["j2"] = chi["d2"].value + "\n\n" + eng["d2"].value
             sta["ab2"].alignment = Alignment(wrap_text = True) #BLE test
             sta["ab2"] = chi["j2"].value + "\n\n" + eng["j2"].value
             sta["ae2"].alignment = Alignment(wrap_text = True) #VM maintainance
             sta["ae2"] = chi["k10"].value + "\n\n" + eng["k10"].value
             sta["ag2"].alignment = Alignment(wrap_text = True) #VM calibrate
             sta["ag2"] = chi["l10"].value + "\n\n" + eng["l10"].value
             sta["ah2"].alignment = Alignment(wrap_text = True) #VM display
             sta["ah2"] = chi["m10"].value + "\n\n" + eng["m10"].value
             sta["aj2"].alignment = Alignment(wrap_text = True) #VM keycodes
             sta["aj2"] = chi["n10"].value + "\n\n" + eng["n10"].value
             sta["aw2"].alignment = Alignment(wrap_text = True) #VM Finish
             sta["aw2"] = chi["aa4"].value + "\n\n" + eng["aa4"].value
         if Value == display[15]: #-------------------------------------------------------OT-fendor
             sta["j2"].alignment = Alignment(wrap_text = True) #BLE setup
             sta["j2"] = chi["d3"].value + "\n\n" + eng["d3"].value
             sta["ab2"].alignment = Alignment(wrap_text = True) #BLE test
             sta["ab2"] = chi["j3"].value + "\n\n" + eng["j3"].value
             sta["ae2"].alignment = Alignment(wrap_text = True) #OT maintainance
             sta["ae2"] = chi["k6"].value + "\n\n" + eng["k6"].value
             sta["ag2"].alignment = Alignment(wrap_text = True) #OT calibrate
             sta["ag2"] = chi["l6"].value + "\n\n" + eng["l6"].value
             sta["ah2"].alignment = Alignment(wrap_text = True) #OT display
             sta["ah2"] = chi["m6"].value + "\n\n" + eng["m6"].value
             sta["aj2"].alignment = Alignment(wrap_text = True) #OT keycodes
             sta["aj2"] = chi["n6"].value + "\n\n" + eng["n6"].value
             sta["aw2"].alignment = Alignment(wrap_text = True) #OT Finish
             sta["aw2"] = chi["aa4"].value + "\n\n" + eng["aa4"].value
         if Value == display[19]:#----------------------------------------------------------Moon
             if Value == "ETVM29818": #---------------------------------------------Special case 29818
                 sta["j2"].alignment = Alignment(wrap_text = True) #BLE setup
                 sta["j2"] = chi["d2"].value + "\n\n" + eng["d2"].value
                 sta["ab2"].alignment = Alignment(wrap_text = True) #BLE test
                 sta["ab2"] = chi["j2"].value + "\n\n" + eng["j2"].value
                 sta["ae2"].alignment = Alignment(wrap_text = True) #VM maintainance
                 sta["ae2"] = eng["k13"].value
                 sta["ag2"].alignment = Alignment(wrap_text = True) #VM calibrate
                 sta["ag2"] = eng["l13"].value
                 sta["ah2"].alignment = Alignment(wrap_text = True) #VM display
                 sta["ah2"] = eng["m13"].value
                 sta["aj2"].alignment = Alignment(wrap_text = True) #VM keycodes
                 sta["aj2"] = eng["n13"].value
                 sta["aw2"].alignment = Alignment(wrap_text = True) #VM Finish
                 sta["aw2"] = eng["aa4"].value
             else:
                 sta["j2"].alignment = Alignment(wrap_text = True) #BLE setup
                 sta["j2"] = chi["d2"].value + "\n\n" + eng["d2"].value
                 sta["ab2"].alignment = Alignment(wrap_text = True) #BLE test
                 sta["ab2"] = chi["j2"].value + "\n\n" + eng["j2"].value
                 sta["ae2"].alignment = Alignment(wrap_text = True) #VM maintainance
                 sta["ae2"] = chi["k10"].value + "\n\n" + eng["k10"].value
                 sta["ag2"].alignment = Alignment(wrap_text = True) #VM calibrate
                 sta["ag2"] = chi["l10"].value + "\n\n" + eng["l10"].value
                 sta["ah2"].alignment = Alignment(wrap_text = True) #VM display
                 sta["ah2"] = chi["m10"].value + "\n\n" + eng["m10"].value
                 sta["aj2"].alignment = Alignment(wrap_text = True) #VM keycodes
                 sta["aj2"] = chi["n10"].value + "\n\n" + eng["n10"].value
                 sta["aw2"].alignment = Alignment(wrap_text = True) #VM Finish
                 sta["aw2"] = chi["aa4"].value + "\n\n" + eng["aa4"].value
                 
            
    
      elif CurrentData == "PowerBoard": #----------------------------------------------------------------------------------------treadmill
         print ("Controller:", Value, "\n")
         for i in range(0,1):
            if Value == controller[i]:
                sta["i2"].alignment = Alignment(wrap_text = True) #Tread 3
                sta["i2"] = chi["c4"].value + "\n\n" + eng["c4"].value
         for i in range(2,20):
            if Value == controller[i]:
                sta["i2"].alignment = Alignment(wrap_text = True) #Fixture Callout
                sta["i2"] = chi["c2"].value + "\n\n" + eng["c2"].value
                sta["z2"].alignment = Alignment(wrap_text = True) #Power up
                sta["z2"] = chi["h2"].value + "\n\n" + eng["h2"].value
                sta["h2"].alignment = Alignment(wrap_text = True) #Equipment needed
                sta["h2"] = chi["b2"].value + "\n" + chi["b7"].value + "\n\n" + eng["b2"].value + "\n" + eng["b7"].value
         if Value == controller[21] or Value == controller[22]:
            sta["i2"].alignment = Alignment(wrap_text = True) #Club fixture Callout
            sta["i2"] = chi["c5"].value + "\n\n" + eng["c5"].value
            sta["z2"].alignment = Alignment(wrap_text = True) #Power up
            sta["z2"] = chi["h5"].value + "\n\n" + eng["h5"].value
         if Value == controller[23]:
            sta["i2"].alignment = Alignment(wrap_text = True) #Rhymebus fixture Callout 
            sta["i2"] = chi["c17"].value + "\n\n" + eng["c17"].value
            sta["z2"].alignment = Alignment(wrap_text = True) #Power up
            sta["z2"] = chi["h18"].value + "\n\n" + eng["h18"].value
         if Value == controller[24]:
            sta["i2"].alignment = Alignment(wrap_text = True) #Delta fixture Callout 
            sta["i2"] = chi["c17"].value + "\n\n" + eng["c17"].value
            sta["z2"].alignment = Alignment(wrap_text = True) #Power up
            sta["z2"] = chi["h18"].value + "\n\n" + eng["h18"].value
         if Value == controller[20] or Value == controller[34]:
            sta["i2"].alignment = Alignment(wrap_text = True) #Boston 1, Boston 2 will require verification of TV or Upright feature
            sta["i2"] = chi["c6"].value + "\n\n" + eng["c6"].value
            sta["z2"].alignment = Alignment(wrap_text = True) #Power up
            sta["z2"] = chi["h2"].value + "\n\n" + eng["h2"].value
         if Value == controller[25]:#--------------------------------------------------------------------------------------------bike and elliptical
            sta["i2"].alignment = Alignment(wrap_text = True) #BASIC_BIKE
            sta["i2"] = chi["c11"].value + "\n\n" + eng["c11"].value
            sta["z2"].alignment = Alignment(wrap_text = True) #Power up
            sta["z2"] = chi["h3"].value + "\n\n" + eng["h3"].value
         if Value == controller[26] or Value == controller[27]:
            sta["i2"].alignment = Alignment(wrap_text = True) #PB_INC
            sta["i2"] = chi["c11"].value + "\n\n" + eng["c11"].value
            sta["z2"].alignment = Alignment(wrap_text = True) #Power up
            sta["z2"] = chi["h3"].value + "\n\n" + eng["h3"].value
         if Value == controller[28] or Value == controller[29]:
            sta["i2"].alignment = Alignment(wrap_text = True) #PB_INC_485
            sta["i2"] = chi["c12"].value + "\n\n" + eng["c12"].value
            sta["z2"].alignment = Alignment(wrap_text = True) #Power up
            sta["Z2"] = chi["h4"].value + "\n\n" + eng["h4"].value
         if Value == controller[30] or Value == controller[31]:
            sta["i2"].alignment = Alignment(wrap_text = True) #PB_PPI_485
            sta["i2"] = chi["c17"].value + "\n\n" + eng["c17"].value
            sta["z2"].alignment = Alignment(wrap_text = True) #Power up
            sta["Z2"] = chi["h17"].value + "\n\n" + eng["h17"].value
         if Value == controller[32] or Value == controller[33]:
            sta["i2"].alignment = Alignment(wrap_text = True) #CLUB_BIKE
            sta["i2"] = chi["c11"].value + "\n\n" + eng["c11"].value
            #sta["z2"].alignment = Alignment(wrap_text = True) #Power up
            #sta["z2"] = chi["h3"].value + "\n\n" + eng["h3"].value
            
      elif CurrentData == "KeyCodeName":#-------------------------------------------------------------------------------------------------
          for i in enumerate(xmlDict['KeyCodeName']):
             if Value[i[0]] == keycode[0]:
                 sta["ae2"].alignment = Alignment(wrap_text = True) #software--------No incline calibrate
                 sta["ae2"] = chi["k9"].value + "\n\n" + eng["k9"].value
                 sta["ah2"].alignment = Alignment(wrap_text = True) #display
                 sta["ah2"] = chi["m9"].value + "\n\n" + eng["m9"].value
                 sta["aj2"].alignment = Alignment(wrap_text = True) #keycodes
                 sta["aj2"] = chi["n8"].value + "\n\n" + eng["n8"].value
             if Value[i[0]] == keycode[0] and keycode[3]:
                 sta["ae2"].alignment = Alignment(wrap_text = True) #software
                 sta["ae2"] = chi["k9"].value + "\n\n" + eng["k9"].value
                 sta["ag2"].alignment = Alignment(wrap_text = True) #incline
                 sta["ag2"] = chi["l8"].value + "\n\n" + eng["l8"].value
                 sta["ah2"].alignment = Alignment(wrap_text = True) #display
                 sta["ah2"] = chi["m9"].value + "\n\n" + eng["m9"].value
                 sta["aj2"].alignment = Alignment(wrap_text = True) #keycodes
                 sta["aj2"] = chi["n8"].value + "\n\n" + eng["n8"].value
             if Value[i[0]] == keycode[1]:
                 sta["ae2"].alignment = Alignment(wrap_text = True) #software--------No incline calibrate
                 sta["ae2"] = chi["k8"].value + "\n\n" + eng["k8"].value
                 sta["ah2"].alignment = Alignment(wrap_text = True) #display
                 sta["ah2"] = chi["m8"].value + "\n\n" + eng["m8"].value
                 sta["aj2"].alignment = Alignment(wrap_text = True) #keycodes
                 sta["aj2"] = chi["n8"].value + "\n\n" + eng["n8"].value
             if Value[i[0]] == keycode[1] and keycode[3]:
                 sta["ae2"].alignment = Alignment(wrap_text = True) #software
                 sta["ae2"] = chi["k8"].value + "\n\n" + eng["k8"].value
                 sta["ag2"].alignment = Alignment(wrap_text = True) #incline
                 sta["ag2"] = chi["l8"].value + "\n\n" + eng["l8"].value
                 sta["ah2"].alignment = Alignment(wrap_text = True) #display
                 sta["ah2"] = chi["m8"].value + "\n\n" + eng["m8"].value
                 sta["aj2"].alignment = Alignment(wrap_text = True) #keycodes
                 sta["aj2"] = chi["n8"].value + "\n\n" + eng["n8"].value
             if Value[i[0]] == keycode[2]:
                 sta["ae2"].alignment = Alignment(wrap_text = True) #software--------No incline calibrate
                 sta["ae2"] = chi["k8"].value + "\n\n" + eng["k8"].value
      elif CurrentData == "ConsoleVoltage":
         print ("Console Voltage:", Value)#---------------------------------------------------------------
         for i in range (1,4):
             if Value == voltage[i]:
                 sta["i2"].alignment = Alignment(wrap_text = True) #Basic Bike1 with voltage > 0; resist and tach and DC power
                 sta["i2"] = chi["c9"].value + "\n\n" + eng["c9"].value
                 sta["z2"].alignment = Alignment(wrap_text = True) #Power up
                 sta["z2"] = chi["h9"].value + "\n\n" + eng["h9"].value
                 sta["am2"].alignment = Alignment(wrap_text = True) #Tach
                 sta["am2"] = chi["p2"].value + "\n\n" + eng["p2"].value
         if Value == voltage[0]:#------------------------------------------------------------------------
             sta["i2"].alignment = Alignment(wrap_text = True) #Basic Bike2 voltage = 0; resist and tach and battery
             sta["i2"] = chi["c10"].value + "\n\n" + eng["c10"].value
             sta["am2"].alignment = Alignment(wrap_text = True) #Tach
             sta["am2"] = chi["p2"].value + "\n\n" + eng["p2"].value
             sta["aw2"].alignment = Alignment(wrap_text = True) #Finish
             sta["aw2"] = chi["aa6"].value + "\n\n" + eng["aa6"].value

      elif CurrentData == "USBHostBoard": #--------------------------------------------------------
          print ("USB:", Value)
          if Value == "true":
              sta["ap2"].alignment = Alignment(wrap_text = True) #USB test-------------------------------------------------------------------test
              sta["ap2"] = chi["s2"].value + "\n\n" + eng["s2"].value
              
      elif CurrentData == "DistanceDriver":
         print ("Tach type:", Value)
         i = 27
         for i in range (27, 30):
             if Value == distance[2] or Value == distance[3] or Value == distance[4] and xmlDict["PowerBoard"] == controller[i]:#----------for acceltach
                 sta["i2"].alignment = Alignment(wrap_text = True)
                 sta["i2"] = chi["c20"].value + "\n\n" + eng["c20"].value
                 sta["z2"].alignment = Alignment(wrap_text = True) #Power up
                 sta["z2"] = chi["h20"].value + "\n\n" + eng["h20"].value
                 sta["am2"].alignment = Alignment(wrap_text = True) #Tach
                 sta["am2"] = chi["p3"].value + "\n\n" + eng["p3"].value
             if Value == distance[0] or Value == distance[1] and xmlDict["PowerBoard"] == controller[i]:#-----------------------------------for lorestach
                 sta["i2"].alignment = Alignment(wrap_text = True)
                 sta["i2"] = chi["c12"].value + "\n\n" + eng["c12"].value
                 sta["z2"].alignment = Alignment(wrap_text = True) #Power up
                 sta["z2"] = chi["h4"].value + "\n\n" + eng["h4"].value
                 sta["am2"].alignment = Alignment(wrap_text = True) #Tach
                 sta["am2"] = chi["p2"].value + "\n\n" + eng["p2"].value
         if "DistanceDriver" in xmlDict and xmlDict['PowerBoard'] == 'BASIC_BIKE':#---------------------------------------------------------for acceltach
             sta["i2"].alignment = Alignment(wrap_text = True)
             sta["i2"] = chi["c19"].value + "\n\n" + eng["c19"].value
             sta["z2"].alignment = Alignment(wrap_text = True) #Power up
             sta["z2"] = chi["h19"].value + "\n\n" + eng["h19"].value
             sta["am2"].alignment = Alignment(wrap_text = True) #Tach
             sta["am2"] = chi["p3"].value + "\n\n" + eng["p3"].value
         if "DistanceDriver" not in xmlDict or xmlDict['DistanceDriver'] == 'LoResTach' and xmlDict['PowerBoard'] == 'BASIC_BIKE':#----------for lorestach
             sta["i2"].alignment = Alignment(wrap_text = True)
             sta["i2"] = chi["c9"].value + "\n\n" + eng["c9"].value
             sta["z2"].alignment = Alignment(wrap_text = True) #Power up
             sta["z2"] = chi["h9"].value + "\n\n" + eng["h9"].value
             sta["am2"].alignment = Alignment(wrap_text = True) #Tach
             sta["am2"] = chi["p2"].value + "\n\n" + eng["p2"].value
#changed to handle array
      elif CurrentData == "PulseDriverItem": #-------------------------------------------------------------------
         for i in enumerate(xmlDict['PulseDriverItem']):
             print ("Pulse Type:", Value[i[0]])
             if Value[i[0]] == pulse[0] or Value[i[0]] == pulse[5]:
                 sta["aq2"].alignment = Alignment(wrap_text = True) #Hand or nanoHand pulse
                 sta["aq2"] = chi["t2"].value + "\n\n" + eng["t2"].value
             if Value[i[0]] == pulse[1]:
                 sta["aq2"].alignment = Alignment(wrap_text = True) #Thumb pulse
                 sta["aq2"] = chi["t3"].value + "\n\n" + eng["t3"].value
             if Value[i[0]] == pulse[2] or Value[i[0]] == pulse[3]:
                 sta["ar2"].alignment = Alignment(wrap_text = True) #Chest pulse
                 sta["ar2"] = chi["t4"].value + "\n\n" + eng["t4"].value
             if Value[i[0]] == pulse[4] or Value[i[0]] == pulse[5]:
                 sta["ar2"].alignment = Alignment(wrap_text = True) #BLE pulse
                 sta["ar2"] = chi["t5"].value + "\n\n" + eng["t5"].value

      elif CurrentData == "FanProtocol":#------------------------------------------------------------------------
         print ("Fan Pin:", Value)
         if Value == fan[0]:
            sta["as2"].alignment = Alignment(wrap_text = True) #Two pin
            sta["as2"] = chi["u2"].value + "\n\n" + eng["u2"].value
         if Value == fan[1]:
            sta["as2"].alignment = Alignment(wrap_text = True) #Three pin
            sta["as2"] = chi["u2"].value + "\n\n" + eng["u2"].value
         if Value == fan[2]:
            sta["as2"].alignment = Alignment(wrap_text = True) #Two and Three pin
            sta["as2"] = chi["u3"].value + "\n\n" + eng["u3"].value
            
#changed to handle array
      elif CurrentData == "AudioSrcItem": #--------------------------------------------------------------------
         for i in enumerate(xmlDict["AudioSrcItem"]):
             print ("Audio Source:", Value[i[0]])
             if Value[i[0]] == audio[1] or Value[i[0]] == audio[2] or Value[i[0]] == audio[3] or Value[i[0]] == audio[4]:
                 sta["at2"].alignment = Alignment(wrap_text = True) #Connect audio source
                 sta["at2"] = chi["v2"].value + "\n\n" + eng["v2"].value
             if Value[i[0]] == audio[5]:
                 sta["at2"].alignment = Alignment(wrap_text = True) #Connect audio source and headphones
                 sta["at2"] = chi["v3"].value + "\n\n" + eng["v3"].value
             if Value[i[0]] == audio[6]:
                 sta["at2"].alignment = Alignment(wrap_text = True) #Connect audio source and headphones and BLE audio
                 sta["at2"] = chi["v4"].value + "\n\n" + eng["v4"].value
             if Value[i[0]] == audio[8]:
                 sta["at2"].alignment = Alignment(wrap_text = True) #Connect MYE audio
                 sta["at2"] = chi["v5"].value + "\n\n" + eng["v5"].value
                  
      elif CurrentData == "TabletProtocol": #-------------------------------------------------------------
          print ("Tablet Type:", Value)
          if Value == tablet[0]:
              sta["w2"].alignment = Alignment(wrap_text = True) #Wolf cosmetics
              sta["w2"] = chi["g3"].value + "\n\n" + eng["g3"].value
              sta["ae2"].alignment = Alignment(wrap_text = True) #Wolf maintainance
              sta["ae2"] = chi["k2"].value + "\n\n" + eng["k2"].value
              sta["ag2"].alignment = Alignment(wrap_text = True) #Wolf calibrate
              sta["ag2"] = chi["l2"].value + "\n\n" + eng["l2"].value
              sta["ah2"].alignment = Alignment(wrap_text = True) #Wolf display blank
              sta["ah2"] = chi["m2"].value + "\n\n" + eng["m2"].value
              sta["aj2"].alignment = Alignment(wrap_text = True) #Wolf key test
              sta["aj2"] = chi["n2"].value + "\n\n" + eng["n2"].value
              sta["ap2"].alignment = Alignment(wrap_text = True) #USB test-------------------------------------------------------------------test
              sta["ap2"] = chi["s2"].value + "\n\n" + eng["s2"].value
              sta["ar2"].alignment = Alignment(wrap_text = True) #Wolf chest pulse
              sta["ar2"] = chi["t6"].value + "\n\n" + eng["t6"].value
              sta["au2"].alignment = Alignment(wrap_text = True) #Wolf HDMI
              sta["au2"] = chi["w2"].value + "\n\n" + eng["w2"].value
              sta["aw2"].alignment = Alignment(wrap_text = True) #Finish
              sta["aw2"] = chi["aa2"].value + "\n\n" + eng["aa2"].value
          if Value == tablet[1]:#------------------------------------------------------------
              sta["ae2"].alignment = Alignment(wrap_text = True) #Legacy maintainance
              sta["ae2"] = chi["k3"].value + "\n\n" + eng["k3"].value
              sta["ag2"].alignment = Alignment(wrap_text = True) #Legacy calibrate
              sta["ag2"] = chi["l3"].value + "\n\n" + eng["l3"].value
              sta["ah2"].alignment = Alignment(wrap_text = True) #Legacy display blank
              sta["ah2"] = chi["m2"].value + "\n\n" + eng["m2"].value
              sta["aj2"].alignment = Alignment(wrap_text = True) #Legacy maintainance
              sta["aj2"] = chi["n3"].value + "\n\n" + eng["n3"].value
              sta["ap2"].alignment = Alignment(wrap_text = True) #USB test-------------------------------------------------------------------test
              sta["ap2"] = chi["s2"].value + "\n\n" + eng["s2"].value
              sta["ar2"].alignment = Alignment(wrap_text = True) #ANT pulse
              sta["ar2"] = chi["t6"].value + "\n\n" + eng["t6"].value
              sta["aw2"].alignment = Alignment(wrap_text = True) #Finish
              sta["aw2"] = chi["aa4"].value + "\n\n" + eng["aa4"].value   
          if Value == tablet[2]:#------------------------------------------------------------
              sta["ae2"].alignment = Alignment(wrap_text = True) #Legacy FreeMotion maintainance
              sta["ae2"] = chi["k4"].value + "\n\n" + eng["k4"].value
              sta["ag2"].alignment = Alignment(wrap_text = True) #Legacy FreeMotion calibrate
              sta["ag2"] = chi["l4"].value + "\n\n" + eng["l4"].value
              sta["ah2"].alignment = Alignment(wrap_text = True) #Legacy FreeMotion display blank
              sta["ah2"] = chi["m2"].value + "\n\n" + eng["m2"].value
              sta["aj2"].alignment = Alignment(wrap_text = True) #Legacy FreeMotion maintainance
              sta["aj2"] = chi["n4"].value + "\n\n" + eng["n4"].value
              sta["ap2"].alignment = Alignment(wrap_text = True) #USB test-------------------------------------------------------------------test
              sta["ap2"] = chi["s2"].value + "\n\n" + eng["s2"].value
              sta["ar2"].alignment = Alignment(wrap_text = True) #ANT pulse
              sta["ar2"] = chi["t4"].value + "\n\n" + eng["t4"].value
              sta["aw2"].alignment = Alignment(wrap_text = True) #Finish
              sta["aw2"] = chi["aa4"].value + "\n\n" + eng["aa4"].value
          if Value == tablet[3]:#------------------------------------------------------------------
              sta["w2"].alignment = Alignment(wrap_text = True) #Argon cosmetics
              sta["w2"] = chi["g3"].value + "\n\n" + eng["g3"].value
              sta["ae2"].alignment = Alignment(wrap_text = True) #Argon maintainance
              sta["ae2"] = chi["k2"].value + "\n\n" + eng["k2"].value
              sta["ag2"].alignment = Alignment(wrap_text = True) #Argon calibrate
              sta["ag2"] = chi["l2"].value + "\n\n" + eng["l2"].value
              sta["ah2"].alignment = Alignment(wrap_text = True) #Argon display blank
              sta["ah2"] = chi["m2"].value + "\n\n" + eng["m2"].value
              sta["aj2"].alignment = Alignment(wrap_text = True) #Argon key test
              sta["aj2"] = chi["n2"].value + "\n\n" + eng["n2"].value
              sta["ar2"].alignment = Alignment(wrap_text = True) #Argon chest pulse
              sta["ar2"] = chi["t6"].value + "\n\n" + eng["t6"].value
              sta["aw2"].alignment = Alignment(wrap_text = True) #Finish
              sta["aw2"] = chi["aa2"].value + "\n\n" + eng["aa2"].value

      elif CurrentData == "ResistanceDriver": #-----------------------------------------------------------
         print ("Resistance Type:",Value)
         sta["an2"].alignment = Alignment(wrap_text = True) #Resistance
         sta["an2"] = chi["r2"].value + "\n\n" + eng["r2"].value
              
      elif CurrentData == "GradeProtocol":
         print ("Incline Type:", Value)
         sta["ao2"].alignment = Alignment(wrap_text = True) #Incline
         sta["ao2"] = chi["q2"].value + "\n\n" + eng["q2"].value
         
      elif CurrentData == "MaintenanceConfigFunction":
         print ("Maintaince Screens: YES")
               

      CurrentData = ""   

      


if ( __name__ == "__main__"):

   # enter the .xml file name including ".xml"
   filename = input("Enter File Name:\n")

   # create an XMLReader
   parser = xml.sax.make_parser()

   # turn off namepsaces
   parser.setFeature(xml.sax.handler.feature_namespaces, 0)

   # override the default ContextHandler
   Handler = Device()
   parser.setContentHandler( Handler )
   parser.parse(filename)


   if ('      ') in pulsearray:
       pulsearray.remove('      ')#--removes last empty value in array
   if ('      ') in audioarray:
       audioarray.remove('      ')#--removes last empty value in array
   
   for i in xmlDict:
       Handler.BuildXml(i,xmlDict[i])


sta["c2"].alignment = Alignment(wrap_text = True) #[PartTypeLink]
sta["c2"] = "控制台 console"

sta["s2"].alignment = Alignment(wrap_text = True) #ESD bag
sta["s2"] = chi["e2"].value + "\n\n" + eng["e2"].value

sta["ay2"].alignment = Alignment(wrap_text = True) #Finish
sta["ay2"] = "PIP_Generator"
sta["ba2"].alignment = Alignment(wrap_text = True) #Finish
sta["ba2"] = str(datetime.now())

mywb.save(sta["e2"].value + "_PIP.xlsx") #save created workbook. Can designate path for saving.

print ("\n")
print ("-------------Procedure Build Is Complete!-------------\n")

j = input ("Create JSON file? Press y or n \n")
if j == 'y':
   #####################
         #JSON#
   #####################
   jsonarray = []
   jsondic = {}
   def FileWriteJson(args):
      tempdic = {}
      with open(sta["e2"].value + ".json", 'w') as file:
          file.write(json.dumps(args, skipkeys= True))
   a = load_workbook(filename = sta["e2"].value + "_PIP.xlsx")
   b = a["Parts - Consoles"]
   for i in b.values:
      jsonarray = i
   for i in enumerate(jsonarray):
      jsondic[header[i[0]+1]] = jsonarray[i[0]]
   FileWriteJson(jsondic)
   #####################
elif j == 'n' or '':
  pass

file = os.path.abspath("" + sta["e2"].value + "_PIP.xlsx")
os.startfile(file)
#os.system("PIP_GEN_id-0.py") #for dev only
os.system("PIP_GEN_id-0.exe") #for release only

input ("\nPress Enter To Close Window:") #for release
